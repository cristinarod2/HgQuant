#!/usr/bin/env python3

import customtkinter as ctk
import pandas as pd
import numpy as np
from datetime import datetime
from tkinter import messagebox, filedialog
import matplotlib.pyplot as plt
from PIL import Image, ImageTk
import io
from datetime import datetime
from dateutil import parser
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font
import os
import sys

if sys.platform == "darwin":
	from AppKit import NSApplication, NSImage
	from Foundation import NSURL

	def set_mac_dock_icon(icon_path="icons/HgQuant.icns"):
		full_path = os.path.abspath(icon_path)
		app = NSApplication.sharedApplication()
		img = NSImage.alloc().initByReferencingFile_(full_path)
		if img and img.size().width > 0:
			app.setApplicationIconImage_(img)	

SIDEBAR_COLOR = "#121A4F"
MAIN_BG_COLOR = "#f4f4f4"
LABEL_TEXT_COLOR = "#1e1e1e"
TEXT ="#445463"




# Bateman calculation functions
def activity_Hg197m(A0, lambda_m, t):
	return A0 * np.exp(-lambda_m * t)

def activity_Hg197g(A0_m, A0_g, lambda_ITm, lambda_m, lambda_g, t):
	term1 = A0_g * np.exp(-lambda_g * t)
	term2 = (lambda_ITm / (lambda_g - lambda_ITm)) * (lambda_g/lambda_m) * A0_m * (np.exp(-lambda_ITm * t) - np.exp(-lambda_g * t))
	return term1 + term2



# Main Application
class RadioactiveDecayApp(ctk.CTk):
	def __init__(self):
		if sys.platform.startswith("win"):
			import ctypes
			ctypes.windll.shell32.SetCurrentProcessExplicitAppUserModelID(u"HgQuant.app")
			
		super().__init__()
		
		if sys.platform == "darwin":
			set_mac_dock_icon()
			
		self.save_path = ""
		self.iconbitmap("icons/HgQuant.ico")  # For Windows
		
		self.protocol("WM_DELETE_WINDOW", self.on_closing)
		
		self.isotope_m = "Hg197m"
		self.isotope_g = "Hg197g"
		self.activity_unit = "KBq"
		
		self.title("HgQuant v.2025.01")
		self.geometry("1000x800")
		ctk.set_appearance_mode("light")
		ctk.set_default_color_theme("blue")
		
		# === Main layout: sidebar + content ===
		main_frame = ctk.CTkFrame(self)
		main_frame.pack(fill="both", expand=True)
		
		# --- Sidebar Layout ---
		sidebar = ctk.CTkFrame(main_frame, width=180, fg_color=SIDEBAR_COLOR)
		sidebar.pack(side="left", fill="y")
		sidebar.pack_propagate(False)
		
		self.icon_calculate = ctk.CTkImage(Image.open("icons/gear.png"), size=(20, 20))
		self.icon_xls = ctk.CTkImage(Image.open("icons/excel.png"), size=(20, 20))
		self.icon_info = ctk.CTkImage(Image.open("icons/info.png"), size=(20, 20))
		self.icon_settings = ctk.CTkImage(Image.open("icons/radio.png"), size=(20, 20))
		
		
		# App icon
		try:
			icon_img = Image.open("icons/HgQuant.png").resize((120, 120))
			self.icon_ctk = ctk.CTkImage(light_image=icon_img, dark_image=icon_img, size=(120, 120))
			icon_label = ctk.CTkLabel(sidebar, image=self.icon_ctk, text="")
			icon_label.pack(pady=(30, 10))
		except Exception as e:
			print(f"⚠️ Could not load icon image: {e}")
			
		# Top section buttons
		ctk.CTkButton(sidebar, text="Calculate & Plot", command=self.calculate_and_plot, fg_color="#3E4A89", hover_color="#5C6BC0", text_color="white", image=self.icon_calculate,compound="left", anchor="w").pack(pady=(10, 5), padx=10, fill="x")
		ctk.CTkButton(sidebar, text="Save Excel", command=self.save_excel, fg_color="#3E4A89", 
			hover_color="#5C6BC0", text_color="white", image=self.icon_xls, compound="left",anchor="w").pack(pady=(10, 5), padx=10, fill="x")
		
		# Bottom section: Info and Settings
		bottom_buttons = ctk.CTkFrame(sidebar, fg_color="transparent")
		bottom_buttons.pack(side="bottom", fill="x", pady=(10, 20))
		
		ctk.CTkButton(bottom_buttons, text="Info", command=self.show_info, fg_color="#3E4A89",       # bluish
			hover_color="#5C6BC0", text_color="white", image=self.icon_info, compound="left",anchor="w").pack(pady=(10, 5), padx=10, fill="x")
		ctk.CTkButton(bottom_buttons, text="Settings", command=self.show_settings, fg_color="#3E4A89",       # bluish
			hover_color="#5C6BC0", text_color="white", image=self.icon_settings, compound="left",anchor="w").pack(pady=(0, 5), padx=10, fill="x")
		
		# === Content Area ===
		self.content_area = ctk.CTkFrame(main_frame, fg_color=MAIN_BG_COLOR)
		self.content_area.pack(side="left", fill="both", expand=True, padx=10, pady=10)
		
		# Title (big and bold)
		self.title_label = ctk.CTkLabel(
			self.content_area,
			text="HgQuant",
			font=ctk.CTkFont(size=20, weight="bold"),
			text_color=TEXT
		)
		self.title_label.pack(pady=(10, 2))
		
		# Subtitle (smaller and lighter)
		self.subtitle_label = ctk.CTkLabel(
			self.content_area,
			text=f"Bateman Decay Calculator {self.isotope_m} → {self.isotope_g}",
			font=ctk.CTkFont(size=14, weight="normal"),
			text_color=TEXT
		)
		self.subtitle_label.pack(pady=(0, 10))
		
		# === Input Frame ===
		input_frame = ctk.CTkFrame(self.content_area, fg_color="#e5e5e5", corner_radius=10)
		input_frame.pack(padx=30, pady=20, fill="x")
		
		# --- Initial Activity Inputs with dynamic labels (after input_frame is defined) ---
		self.label_activity_g = ctk.CTkLabel(input_frame, text=f"Initial {self.isotope_g} ({self.activity_unit}):")
		self.label_activity_g.grid(row=0, column=0, padx=5, pady=5, sticky="e")
		
		self.hg197g_initial = ctk.CTkEntry(input_frame, border_width=1)
		self.hg197g_initial.grid(row=0, column=1, padx=5, pady=5, sticky="ew")
		
		self.label_activity_m = ctk.CTkLabel(input_frame, text=f"Initial {self.isotope_m} ({self.activity_unit}):")
		self.label_activity_m.grid(row=1, column=0, padx=5, pady=5, sticky="e")
		
		self.hg197m_initial = ctk.CTkEntry(input_frame, border_width=1)
		self.hg197m_initial.grid(row=1, column=1, padx=5, pady=5, sticky="ew")
		
		
		# Half-life
		self.label_halflife_g = ctk.CTkLabel(input_frame, text=f"Half-life {self.isotope_g} (h):")
		self.label_halflife_g.grid(row=0, column=2, padx=5, pady=5, sticky="e")
		self.hg197g_halflife = ctk.CTkEntry(input_frame, border_width=1, fg_color="#fafa87")
		self.hg197g_halflife.grid(row=0, column=3, padx=5, pady=5)
		self.hg197g_halflife.insert(0, "64.14")  # Auto-filled half-life of Hg-197g
		
		self.label_halflife_m = ctk.CTkLabel(input_frame, text=f"Half-life {self.isotope_m} (h):")
		self.label_halflife_m.grid(row=1, column=2, padx=5, pady=5)
		self.hg197m_halflife = ctk.CTkEntry(input_frame, border_width=1, fg_color="#fafa87")
		self.hg197m_halflife.grid(row=1, column=3, padx=5, pady=5)
		self.hg197m_halflife.insert(0, "23.8")  # Auto-filled half-life of Hg-197m
		
		# Initial Datetime
		ctk.CTkLabel(input_frame, text="Measurement Timestamp:").grid(row=2, column=0, padx=5, pady=5, sticky="e")
		self.initial_datetime = ctk.CTkEntry(input_frame, border_width=1)
		self.initial_datetime.grid(row=2, column=1, columnspan=3, padx=10, pady=6, sticky="ew")
		
		self.datetime_hint = "YYYY-MM-DD HH:MM[:SS]"
		self.initial_datetime.insert(0, self.datetime_hint)
		self.initial_datetime.configure(text_color="gray")
		
		def clear_datetime_hint(event):
			if not self.winfo_exists():
				return
			if self.initial_datetime.get().strip() == self.datetime_hint:
				self.initial_datetime.delete(0, "end")
				self.initial_datetime.configure(text_color="black")
				
		def restore_datetime_hint(event):
			if not self.winfo_exists():
				return
			if not self.initial_datetime.get().strip():
				self.initial_datetime.insert(0, self.datetime_hint)
				self.initial_datetime.configure(text_color="gray")
				
		self.initial_datetime.bind("<FocusIn>", clear_datetime_hint)
		self.initial_datetime.bind("<FocusOut>", restore_datetime_hint)
		
		# Timepoints input
		ctk.CTkLabel(input_frame, text="Timepoints (one per line):").grid(row=3, column=0, padx=5, pady=5)
		self.timepoints_text = ctk.CTkTextbox(input_frame, border_width=1, height=150, width=300)
		self.timepoints_text.grid(row=3, column=1, columnspan=3, padx=5, pady=5, sticky="ew")
		
		
				
	def on_closing(self):
		plt.close('all')  # close any matplotlib windows
		
		# Unbind any lingering focus callbacks
		try:
			self.initial_datetime.unbind("<FocusIn>")
			self.initial_datetime.unbind("<FocusOut>")
		except:
			pass
			
		self.destroy()		
	
		
	def calculate_and_plot(self):
		try:
			A_Hg197g_0 = float(self.hg197g_initial.get())
			A_Hg197m_0 = float(self.hg197m_initial.get())
			T_half_g = float(self.hg197g_halflife.get())
			T_half_m = float(self.hg197m_halflife.get())
			# Clean and parse the initial time
			# Get user input or use default
			initial_raw = self.initial_datetime.get().strip()
			
			self.am0 = A_Hg197m_0
			self.ag0 = A_Hg197g_0
			self.t_half_m = T_half_m
			self.t_half_g = T_half_g
			self.measured_time = parser.parse(initial_raw)
		
			
			if not initial_raw:
				initial_raw = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
				self.initial_datetime.insert(0, initial_raw)  
								
			try:
				initial_dt = parser.parse(initial_raw)
			except Exception:
				messagebox.showerror("Date Error", "Please enter a valid date in the format YYYY-MM-DD or YYYY-MM-DD HH:MM[:SS]")
				return
			
				# Get and clean all timepoints
			time_points_raw = self.timepoints_text.get("1.0", "end").strip().split("\n")
			time_points = [tp.strip().split('.')[0] for tp in time_points_raw if tp.strip()]			
			
			self.lambda_g = round(np.log(2) / T_half_g, 8)
			self.lambda_m = round(np.log(2) / T_half_m, 8)
			self.lambda_ITm = round(0.914 * self.lambda_m, 8)
			lambda_decay_to_g = self.lambda_ITm
			
			# Parse timepoints and truncate to minute precision
			time_points_dt = []
			for tp in time_points:
				try:
					time_points_dt.append(parser.parse(tp.strip()))
				except Exception:
					messagebox.showerror("Date Error", f"Invalid timepoint:\n{tp}")
					return			
			# Compute elapsed time in hours (same formula as Excel)
			time_elapsed_h = [(tp - initial_dt).total_seconds() / 3600 for tp in time_points_dt]
			
			hg197m = np.array([activity_Hg197m(A_Hg197m_0, self.lambda_m, t) for t in time_elapsed_h])
			hg197g = np.array([
				activity_Hg197g(A_Hg197m_0, A_Hg197g_0, lambda_decay_to_g, self.lambda_m, self.lambda_g, t)
				for t in time_elapsed_h
			])
			
			total_activity = hg197m + hg197g
			hg197m_pct = hg197m / total_activity * 100
			hg197g_pct = hg197g / total_activity * 100
			
			self.activities_df = pd.DataFrame({
				"Time Point": time_points,
				"Hours Elapsed": time_elapsed_h,
				f"{self.isotope_m} ({self.activity_unit})": hg197m,
				f"{self.isotope_g} ({self.activity_unit})": hg197g,
				f"% {self.isotope_m}": hg197m_pct,
				f"% {self.isotope_g}": hg197g_pct
			})
			self.activities_df["Decay Factor Hg-197m"] = hg197m / A_Hg197m_0
			self.activities_df["Decay Factor Hg-197g"] = hg197g / (A_Hg197m_0 + A_Hg197g_0)
			
			# Plot activities
			plt.figure(figsize=(10, 6))
			plt.plot(time_elapsed_h, hg197m, 'o-', label=self.isotope_m)
			plt.plot(time_elapsed_h, hg197g, 's-', label=self.isotope_g)
			plt.xlabel('Hours Elapsed')
			plt.ylabel(f'Activity ({self.activity_unit})')
			plt.title('Radioactive Decay')
			plt.grid()
			plt.legend()
			plt.show()
			
			
			# Plot percentage over time
			plt.figure(figsize=(10, 6))
			plt.plot(time_elapsed_h, hg197m_pct, 'o-', label=f"% {self.isotope_m}")
			plt.plot(time_elapsed_h, hg197g_pct, 's-', label=f"% {self.isotope_g}")
			plt.xlabel('Hours Elapsed')
			plt.ylabel('Percentage (%)')
			plt.title(f'Percentage of {self.isotope_m} and {self.isotope_g} Over Time')
			plt.grid()
			plt.legend()
			plt.show()
			
			
			# Clear old results section
			if hasattr(self, 'results_container'):
				self.results_container.destroy()
			self.results_container = ctk.CTkFrame(self)
			self.results_container.pack(pady=(10, 15), fill="both", expand=True)
			
			# LEFT side (constants + Bateman equations)
			left_frame = ctk.CTkFrame(self.results_container, fg_color="transparent")
			left_frame.pack(side="left", fill="y", padx=10)
			
			self.constants_table_frame = ctk.CTkFrame(left_frame, fg_color="transparent")
			self.constants_table_frame.pack(pady=(0, 10))
			
			headers = ["Decay Constant", "Value (h⁻¹)"]
			values = [
				[f"λm ({self.isotope_m})", f"{self.lambda_m:.8f}"],
				[f"λ_ITm ({self.isotope_m} → {self.isotope_g})", f"{self.lambda_ITm:.8f}"],
				[f"λg ({self.isotope_g})", f"{self.lambda_g:.8f}"]
			]
			
			for col, header in enumerate(headers):
				ctk.CTkLabel(self.constants_table_frame, text=header, font=("Helvetica", 13, "bold")).grid(row=0, column=col, padx=10, pady=4)
				
			for row, (name, val) in enumerate(values, start=1):
				ctk.CTkLabel(self.constants_table_frame, fg_color ="transparent", text=name, font=("Helvetica", 13)).grid(row=row, column=0, padx=10, pady=2)
				ctk.CTkLabel(self.constants_table_frame, text=val, font=("Helvetica", 13)).grid(row=row, column=1, padx=10, pady=2)
				
			bateman_img = self.render_bateman_equation_image()
			self.bateman_img_ctk = ctk.CTkImage(light_image=bateman_img, dark_image=bateman_img, size=(300, 130))
			
			# Frame to wrap image and title
			equation_frame = ctk.CTkFrame(left_frame, fg_color="transparent")
			equation_frame.pack(pady=(15, 10), padx=5)
			
			# Optional title above the equation
			ctk.CTkLabel(equation_frame, text="Bateman Equations", font=("Helvetica", 13, "italic")).pack(pady=(0, 5))
			
			# The equation image
			self.bateman_label = ctk.CTkLabel(equation_frame, image=self.bateman_img_ctk, text="")
			self.bateman_label.pack()
			
			# RIGHT side (results table)
			right_frame = ctk.CTkFrame(self.results_container)
			right_frame.pack(side="left", fill="both", expand=True, padx=10)
			
			if hasattr(self, 'results_textbox'):
				self.results_textbox.destroy()
			self.results_textbox = ctk.CTkTextbox(right_frame, width=400, height=220, font=("Courier New", 11), wrap="none")
			self.results_textbox.pack(fill="both", expand=True)
			
			# Format each row with aligned, padded columns
			m_col = f"{self.isotope_m} ({self.activity_unit})"
			g_col = f"{self.isotope_g} ({self.activity_unit})"
			m_pct_col = f"% {self.isotope_m}"
			g_pct_col = f"% {self.isotope_g}"
			
			
			# Insert results into textbox
			# Format and align columns manually for readability
			table_lines = []
						# Header row with fixed-width column names
			df_m = self.activities_df["Decay Factor Hg-197m"]
			df_g = self.activities_df["Decay Factor Hg-197g"]
			
			header = f"{'Time Point':<20} {'Elapsed (h)':>12} {m_col:>15} {g_col:>15} {m_pct_col:>12} {g_pct_col:>12} {'DF m':>10} {'DF g':>10}"
			table_lines.append(header)
			table_lines.append("-" * len(header))
			
			for i, row in self.activities_df.iterrows():
				line = (
					f"{row['Time Point']:<20} "
					f"{row['Hours Elapsed']:>12.3f} "
					f"{row[m_col]:>15.3f} "
					f"{row[g_col]:>15.3f} "
					f"{row[m_pct_col]:>12.2f} "
					f"{row[g_pct_col]:>12.2f} "
					f"{row['Decay Factor Hg-197m']:>10.4f} "
					f"{row['Decay Factor Hg-197g']:>10.4f}"
				)
				table_lines.append(line)				
			# Join all lines into one string
			table_str = "\n".join(table_lines)
			self.results_textbox.insert("1.0", table_str)
			self.results_textbox.configure(state="disabled", wrap="none")
			
			
		except Exception as e:
			messagebox.showerror("Error", f"An error occurred:\n{e}")
			
			
	def render_bateman_equation_image(self):
		import matplotlib.pyplot as plt
		from matplotlib import rcParams
		
		rcParams.update({
			"text.usetex": False,
			"font.size": 16,
			"mathtext.fontset": "stix",  # or "cm"
		})
		
		equation_text = (
			r"$A_m(t) = A_{m0} \cdot e^{-\lambda_m t}$" "\n\n"
			r"$A_g(t) = A_{g0} \cdot e^{-\lambda_g t} + \left("
			r"\frac{\lambda_{ITm}}{\lambda_g - \lambda_{ITm}} \cdot \frac{\lambda_g}{\lambda_m} \cdot A_{m0} \cdot "
			r"\left(e^{-\lambda_{ITm} t} - e^{-\lambda_g t}\right)"
			r"\right)$"
		)
		
		fig, ax = plt.subplots(figsize=(6, 2))
		ax.text(0.5, 0.5, equation_text, fontsize=16, ha='center', va='center')
		ax.axis("off")
		
		buf = io.BytesIO()
		plt.savefig(buf, format='png', dpi=250, bbox_inches='tight', transparent=True)
		buf.seek(0)
		plt.close(fig)
		
		return Image.open(buf)
	
			
	def save_excel(self):
		import os
		import pandas as pd
		from tkinter import filedialog, messagebox
		from datetime import datetime
		from openpyxl.styles import Font
		
		if not hasattr(self, 'activities_df'):
			messagebox.showwarning("Warning", "Please run 'Calculate & Plot' before saving.")
			return
		
		if self.save_path:
			default_filename = f"BatemanDecay_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
			filepath = os.path.join(self.save_path, default_filename)
		else:
			filepath = filedialog.asksaveasfilename(defaultextension=".xlsx", filetypes=[("Excel files", "*.xlsx")])
			
		if not filepath:
			return
		
		required_attrs = ['am0', 'ag0', 't_half_m', 't_half_g', 'measured_time', 'lambda_m', 'lambda_ITm', 'lambda_g']
		if not all(hasattr(self, attr) for attr in required_attrs):
			messagebox.showwarning("Missing Data", "Please run 'Calculate & Plot' before saving.")
			return
		
		# === Prepare DataFrames ===
		df_metadata = pd.DataFrame([
			["Report generated by", "HgQuant v.2025.01"],
			["on", datetime.now().strftime("%Y-%m-%d %H:%M:%S")],
			["User-defined Isotopes", f"{self.isotope_m} → {self.isotope_g}"],
			["GitHub", "https://github.com/cristinarod2/HgQuant"]
		], columns=["Report", "Value"])
		
		df_inputs = pd.DataFrame({
			"Parameter": [
				f"Initial Activity ({self.isotope_m})",
				f"Initial Activity ({self.isotope_g})",
				f"Half-life ({self.isotope_m})",
				f"Half-life ({self.isotope_g})",
				"Activity Units",
				"Measured Timestamp"
			],
			"Value": [
				self.am0,
				self.ag0,
				f"{self.t_half_m:.4f} h",
				f"{self.t_half_g:.4f} h",
				self.activity_unit,
				self.measured_time.strftime("%Y-%m-%d %H:%M:%S")
			]
		})		
		df_constants = pd.DataFrame({
			"Decay Constant": ["λm (Hg-197m)", "λ_ITm (Hg-197m → Hg-197g)", "λg (Hg-197g)"],
			"Value (h⁻¹)": [
				f"{self.lambda_m:.8f}",
				f"{self.lambda_ITm:.8f}",
				f"{self.lambda_g:.8f}"
			]
		})
		
		# === Write to Excel ===
		try:
			with pd.ExcelWriter(filepath, engine="openpyxl") as writer:
				sheet = "Input and Decay Constants"
				
				# Write all dataframes with spacing
				df_metadata.to_excel(writer, index=False, sheet_name=sheet, startrow=0)
				df_inputs.to_excel(writer, index=False, sheet_name=sheet, startrow=len(df_metadata) + 2)
				df_constants.to_excel(writer, index=False, sheet_name=sheet, startrow=len(df_metadata) + len(df_inputs) + 5)
				self.activities_df.to_excel(writer, index=False, sheet_name="Decay Table")
				
				ws = writer.sheets[sheet]
				
				# === Style: Fonts ===
				font_regular = Font(name="Calibri", size=12)
				font_bold = Font(name="Calibri", size=16, bold=True)
				
				# Apply font to metadata table
				for row in ws.iter_rows(min_row=1, max_row=len(df_metadata), min_col=1, max_col=2):
					for cell in row:
						cell.font = font_regular
						
				for cell in ws[f"A1:B1"]:
					for c in cell:
						c.font = font_bold
					
				# Bold parameter + constant headers
				inputs_start = len(df_metadata) + 3
				constants_start = len(df_metadata) + len(df_inputs) + 6
				
				for cell in ws[f"A{inputs_start}:B{inputs_start}"][0]:
					cell.font = font_bold
					
				for cell in ws[f"A{constants_start}:B{constants_start}"][0]:
					cell.font = font_bold
					
				# Hyperlink the GitHub row in metadata
				ws["B5"].hyperlink = "https://github.com/cristinarod2/HgQuant"
				ws["B5"].style = "Hyperlink"
				
				# Optional: Freeze top row to keep metadata visible
				#ws.freeze_panes = f"A{len(df_metadata) + 2}"
				
				# Auto-adjust column widths
				for sheetname in writer.sheets:
					ws_auto = writer.sheets[sheetname]
					if sheetname == sheet:
						dfs_to_check = [df_metadata, df_inputs, df_constants]
					else:
						dfs_to_check = [self.activities_df]
						
					for df in dfs_to_check:
						for i, column in enumerate(df.columns):
							col_letter = get_column_letter(i + 1)
							column_name = str(df.columns[i])
							
							try:
								content_len = df[column].astype(str).map(len).max()
							except Exception:
								content_len = 0  # fallback if column data is missing or empty
								
							header_len = len(column_name)
							max_len = max(header_len, content_len) + 4
							ws_auto.column_dimensions[col_letter].width = max_len
							
			messagebox.showinfo("Saved", f"Excel file saved successfully:\n{filepath}")
			
		except Exception as e:
			messagebox.showerror("Save Error", f"Could not save Excel file:\n{e}")
			
	def show_info(self):
		info_win = ctk.CTkToplevel(self)
		info_win.title("About HgQuant")
		info_win.geometry("600x750")
		info_win.resizable(False, False)
		
		# Load and set the icon
		try:
			icon_img = Image.open("icons/HgQuant.png").resize((120, 120))
			self.icon_ctk = ctk.CTkImage(light_image=icon_img, dark_image=icon_img, size=(120, 120))
			icon_label = ctk.CTkLabel(info_win, image=self.icon_ctk, text="")
			icon_label.pack(pady=(25, 10))
		except Exception as e:
			print(f"⚠️ Could not load HgQuant icon: {e}")
			
		# App title
		title_label = ctk.CTkLabel(
			info_win,
			text="HgQuant – Bateman Decay Calculator",
			font=ctk.CTkFont(size=17, weight="bold"),
			text_color=TEXT
		)
		title_label.pack(pady=(0, 8))
		
		# Version and author
		meta_label = ctk.CTkLabel(
			info_win,
			text=(
				"v.2025.01\n"
				"Cristina Rodriguez, PhD\n"
				"University of British Columbia – 2025\n\n"
				"Built with Python 3.11, Matplotlib, and CustomTkinter\n"
				"Developed using open-source tools for research and teaching"
			),
			font=ctk.CTkFont(size=12),
			text_color=TEXT,
			justify="center"
		)
		meta_label.pack(padx=20, pady=(0, 15))
		
		# Divider
		divider = ctk.CTkLabel(
			info_win,
			text="─" * 42,
			font=ctk.CTkFont(size=12),
			text_color="#999999"
		)
		divider.pack(pady=(0, 10))
		
		# Description
		description = (
			"HgQuant is a focused tool for modeling the decay of Mercury-197 isomers\n "
			"(¹⁹⁷ᵐHg → ¹⁹⁷gHg) using Bateman equations.\n\n"
			"Designed specifically for dual-isotope SPECT workflows, HgQuant simulates time-resolved activity "
			"of ¹⁹⁷ᵐHg and ¹⁹⁷gHg, incorporating both independent decay and internal transition.\n\n"
			"Key features include:\n"
			"• Timepoint-based modeling of ¹⁹⁷ᵐHg/¹⁹⁷gHg decay\n"
			"• Automatic calculation of buildup and effective decay\n"
			"• Customizable decay constants for phantom or in vivo datasets\n"
			"• Excel exports and clean, publication-ready plots\n\n"
			"HgQuant streamlines decay correction and quantification for Hg-197 imaging and biodistribution studies, "
			"improving accuracy and reproducibility."
		)
		desc_label = ctk.CTkLabel(
			info_win,
			text=description,
			wraplength=500,
			justify="center",
			font=ctk.CTkFont(size=13),
			text_color=TEXT
		)
		desc_label.pack(padx=20)
		
		# Close button
		ctk.CTkButton(info_win, text="Close", command=info_win.destroy).pack(pady=(15, 10))
		
	def show_settings(self):
		settings_window = ctk.CTkToplevel(self)
		settings_window.title("Settings")
		settings_window.geometry("300x500")
		
		ctk.CTkLabel(settings_window, text="Folder to Save Files", font=("Helvetica", 14, "bold")).pack(pady=(10, 5))
		self.path_display = ctk.CTkLabel(settings_window, text=self.save_path or "No folder selected", wraplength=480)
		self.path_display.pack(pady=(0, 10))
		
		def choose_folder():
			folder = filedialog.askdirectory()
			if folder:
				self.save_path = folder
				self.path_display.configure(text=self.save_path)
				
		ctk.CTkButton(settings_window, text="Choose Folder", command=choose_folder).pack(pady=5)
		
		# --- Isotope Settings ---
		ctk.CTkLabel(settings_window, text="Isotope Names").pack(pady=(15, 2))
		isotope_frame = ctk.CTkFrame(settings_window)
		isotope_frame.pack(pady=5)
		
		ctk.CTkLabel(isotope_frame, text="Parent (m):").grid(row=0, column=0, padx=5, pady=5)
		isotope_m_entry = ctk.CTkEntry(isotope_frame)
		isotope_m_entry.insert(0, self.isotope_m)
		isotope_m_entry.grid(row=0, column=1)
		
		ctk.CTkLabel(isotope_frame, text="Daughter (g):").grid(row=1, column=0, padx=5, pady=5)
		isotope_g_entry = ctk.CTkEntry(isotope_frame)
		isotope_g_entry.insert(0, self.isotope_g)
		isotope_g_entry.grid(row=1, column=1)
		
		# --- Activity Unit ---
		ctk.CTkLabel(settings_window, text="Activity Unit").pack(pady=(10, 2))
		unit_option = ctk.CTkOptionMenu(settings_window, values=["KBq", "MBq", "µCi", "KBq/mL", "MBq/mL", "µCi/mL"])
		unit_option.set(self.activity_unit)
		unit_option.pack()
		
		# --- Half-Life Settings ---
		ctk.CTkLabel(settings_window, text="Half-lives (in hours)").pack(pady=(10, 2))
		halflife_frame = ctk.CTkFrame(settings_window)
		halflife_frame.pack(pady=5)
		
		ctk.CTkLabel(halflife_frame, text=f"{self.isotope_m}:").grid(row=0, column=0, padx=5, pady=5)
		halflife_m_entry = ctk.CTkEntry(halflife_frame)
		halflife_m_entry.insert(0, self.hg197m_halflife.get())
		halflife_m_entry.grid(row=0, column=1)
		
		ctk.CTkLabel(halflife_frame, text=f"{self.isotope_g}:").grid(row=1, column=0, padx=5, pady=5)
		halflife_g_entry = ctk.CTkEntry(halflife_frame)
		halflife_g_entry.insert(0, self.hg197g_halflife.get())
		halflife_g_entry.grid(row=1, column=1)
		
		
		def save_settings():
			self.isotope_m = isotope_m_entry.get()
			self.isotope_g = isotope_g_entry.get()
			self.activity_unit = unit_option.get()
			
			# Update half-life entries BEFORE destroying the window
			self.hg197m_halflife.delete(0, "end")
			self.hg197m_halflife.insert(0, halflife_m_entry.get())
			
			self.hg197g_halflife.delete(0, "end")
			self.hg197g_halflife.insert(0, halflife_g_entry.get())
			
			self.update_labels()
			settings_window.destroy()
			
			
		ctk.CTkButton(settings_window, text="Save", command=save_settings).pack(pady=(15, 10))
		
	@staticmethod
	def convert_activity(value, from_unit, to_unit):
		conversion_factors = {
			("KBq/mL", "MBq/mL"): 0.001,
			("KBq/mL", "µCi/mL"): 0.02703,
			("MBq/mL", "KBq/mL"): 1000,
			("MBq/mL", "µCi/mL"): 27.03,
			("µCi/mL", "KBq/mL"): 37.04,
			("µCi/mL", "MBq/mL"): 0.03704,
		}
		if from_unit == to_unit:
			return value
		factor = conversion_factors.get((from_unit, to_unit))
		if factor is None:
			raise ValueError(f"Unsupported conversion: {from_unit} → {to_unit}")
		return value * factor
	
	def update_labels(self):
		self.title(f"Bateman Decay Calculator: {self.isotope_m} → {self.isotope_g}")
		
		# Update dynamic input labels if they exist
		if hasattr(self, 'title_label'):
			self.title_label.configure(text=f"Bateman Decay Calculator: {self.isotope_m} → {self.isotope_g}")
			
		if hasattr(self, 'label_activity_m'):
			self.label_activity_m.configure(text=f"Initial {self.isotope_m} ({self.activity_unit}):")
		if hasattr(self, 'label_activity_g'):
			self.label_activity_g.configure(text=f"Initial {self.isotope_g} ({self.activity_unit}):")
			
		if hasattr(self, 'label_halflife_m'):
			self.label_halflife_m.configure(text=f"Half-life {self.isotope_m} (h):")
		if hasattr(self, 'label_halflife_g'):
			self.label_halflife_g.configure(text=f"Half-life {self.isotope_g} (h):")
			
			
		# Update Bateman equation text if exists
		if hasattr(self, 'bateman_label'):
			self.bateman_label.configure(text=(
				f"Bateman Equations:\n\n"
				f"{self.isotope_m}:\n"
				"Aₘ(t) = Aₘ₀ · e^(−λₘ·t)\n\n"
				f"{self.isotope_g}:\n"
				"A_g(t) = A_g₀ · e^(−λg·t) + "
				"(λ_ITm / (λg − λₘ)) · Aₘ₀ · (e^(−λₘ·t) − e^(−λg·t))"
			))
			
			
			
if __name__ == "__main__":
	app = RadioactiveDecayApp()
	app.mainloop()
	